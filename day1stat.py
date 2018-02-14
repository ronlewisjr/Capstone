import os
import cx_Oracle
import csv
import pandas as pd
import numpy as np
import datetime
import win32com.client
from openpyxl import load_workbook
from openpyxl import Workbook
import win32com.client
from datetime import datetime, date, timedelta

con = cx_Oracle.connect('pricing/thepriceisright@SHDATA')
cur = con.cursor()

# create query and list with all unique reps or unique bid owners
cur.execute("""
select
tab1.bid_owner,
tab2.emp_name,
tab2.modified_EMAIL,
count(*) as tot_cnt3
from
(
select distinct bid_owner from
(
select
bid_owner,
count(*) as tot_cnt
FROM KGEK_MASTER_TABLE_V3
where sell_price >= weighted_cost and sell_price>=COMP_COST --added 20171229
group by bid_owner
order by bid_owner asc
)
) tab1
inner join
(
select
tab_email.emp_name,
tab_email.modified_EMAIL,
count(*)  as tot_cnt2
from
(
select
tab_main.emp_name,
case
when tab_main.emp_name in('BRENT KIEFER') then 'BRENT.KIEFER@MCKESSON.COM'
when tab_main.emp_name in('BRENT POYTHRESS') then 'Brent.Poythress@McKesson.com'
when tab_main.emp_name in('BRIAN JONES') then 'Brian.Jones@McKesson.com'
when tab_main.emp_name in('JANICE BAYNAI') then 'Janice.Baynai@McKesson.com'
when tab_main.emp_name in('JANINE GIANATASIO') then 'JANINE.GIANATASIO@MCKESSON.COM'
when tab_main.emp_name in('JEFF SOUTHARD') then 'JEFF.SOUTHARD@MCKESSON.COM'
when tab_main.emp_name in('JEFF WARD') then 'JEFF.WARD@MCKESSON.COM'
when tab_main.emp_name in('JIM KOHLER') then 'JIM.KOHLER@MCKESSON.COM'
when tab_main.emp_name in('JIM SMELLEY') then 'Jim.Smelley@mckesson.com'
when tab_main.emp_name in('JEFF WARD') then 'JEFF.WARD@MCKESSON.COM'
when tab_main.emp_name in('LINDSEY SHADDUCK OPEN') then 'LINDSEY.SHADDUCK@MCKESSON.COM'
when tab_main.emp_name in('MARK SNODGRASS') then 'Mark.Snodgrass@McKesson.com'
when tab_main.emp_name in('MAURINE CAVANAUGH') then 'Maurine.Cavanaugh@McKesson.com'
when tab_main.emp_name in('MICHAEL CARNI') then 'Michael.Carni@McKesson.com'
when tab_main.emp_name in('MICHAEL KENNEDY') then 'MICHAEL.KENNEDY@MCKESSON.COM'
when tab_main.emp_name in('MONICA JOHNSON') then 'MONICA.JOHNSON@MCKESSON.COM'
when tab_main.emp_name in('MONICA JOHNSON OPEN') then 'MONICA.JOHNSON@MCKESSON.COM'
when tab_main.emp_name in('NOBLE FLEMING') then 'NOBLE.FLEMING@MCKESSON.COM'
when tab_main.emp_name in('OPEN BRIAN SLUSAW') then 'BRIAN.SLUSAW@MCKESSON.COM'
when tab_main.emp_name in('OPEN LORI LIKENS') then 'LORI.LIKENS@MCKESSON.COM'
when tab_main.emp_name in('PHIL BEVACQUA') then 'PHIL.BEVACQUA@MCKESSON.COM'
when tab_main.emp_name in('RAMIRO MOLINA') then 'Ramiro.Molina@McKesson.com'
when tab_main.emp_name in('ROBERT PETERSEN') then 'Robert.Petersen@McKesson.com'
when tab_main.emp_name in('RUTH SCHULTZ') then 'RUTH.SCHULTZ@MCKESSON.COM'
when tab_main.emp_name in('TESS SIMMS') then 'TESS.SIMMS@MCKESSON.COM'
else tab_main.E_MAIL
end as modified_EMAIL
FROM KGEK_BID_OWNER  tab_main
) tab_email
group by emp_name, modified_EMAIL
) tab2
on tab1.bid_owner = tab2.emp_name
where BID_OWNER not in ('CHAIN ACCOUNTS', 'HOUSE ACCOUNTS', 'OPEN TERR 113', 'OPEN TERR 118', 'OPEN TERR 129', 'OPEN TERR 133', 'OPEN TERR 143', 'OPEN TERR 731 SAM', 'LINDSEY SHADDUCK OPEN', 'MONICA JOHNSON OPEN', 'OPEN BRIAN SLUSAW', 'OPEN LORI LIKENS')
group by tab1.bid_owner,
tab2.emp_name,
modified_EMAIL
order by tab1.bid_owner asc
""")
bid_owner_list = cur.fetchall()
count_total_bid_owners = len(bid_owner_list)

cur.execute("""
select
account_rep,
count(*) as tot_accts
from PRICING.EHVXZ0J_HF_FINAL_DRIVER_011218
where account_rep not in(
'OPEN TERR 115',
'OPEN TERR 129',
'OPEN TERR 234',
'OPEN TERR 854',
'OPEN TERR 859',
'OPEN TERR 868',
'OPEN - OH'
) and account_rep is not null
group by account_rep
order by account_rep asc
""")
HF_account_rep_list = cur.fetchall()


#count_total_bid_owners = 186 # added to cut down on number of reports
current_bid_owner = 0
list_of_file_locations = []
current_date = time = str(datetime.now().date())

list_of_bid_owners_with_HF_tab = []
list_of_bid_owners_without_HF_tab = []

while current_bid_owner <  count_total_bid_owners:  #2:
    # load template file and create variables for each sheet
    wb = load_workbook(filename = 'C:\\Users\\ee1nj1k\\Documents\\Margin_Day\\201802\\WIP_Report\\Template_HF_20180115.xlsm', read_only=False, keep_vba=True)
    #wb = load_workbook(filename = 'C:\\Users\\ee1nj1k\\Documents\\Python\\Python_Scripts\\Margin_Day_June_2017\\Template_201706_FINAL_edits_20170525.xlsm', read_only=False, keep_vba=True)
    tab1 = wb['Step1-CustomerAction']
    tab2 = wb['Step2-Positive Margin Items']
    tab3 = wb['Grouping Breakout']
    tab4 = wb['Formatting']
    tab5 = wb['Notes']
    tab_negative_margin_items = wb['Step3-Negative Margin Items']
    tab_HF = wb['Step4-Handling Fee Increases']
    # take the next bid owner name from the list
    owner_name1, emp_name, bid_owner_email, ignore_cnt = bid_owner_list[current_bid_owner]
    # query the main table to obtain bid owner-level input data
    tab1_query = f"select prcsrc, bhdesc, .04 as Options, bid_owner, count(*) as tot_cnt from (select prcsrc, bhdesc, .04 as Options, bid_owner from KGEK_MASTER_TABLE_V3 where bid_owner in(\'{owner_name1}\')) tab1 group by prcsrc, bhdesc, Options, bid_owner order by prcsrc desc"
    tab2_query = f"select t1.SVP_name, t1.Rep_mgr_name, t1.bid_owner, t1.prcsrc, t1.ACCT, t1.Account_name, t1.ITEM_DESCRIPTION, t1.MFG_CODE, t1.prod_cat_desc, t1.RLITEM, t1.QTY_6MO, cast(t1.SALES_6MO as decimal(18,2)) as sales_6mo6, t1.BHTYPE, cast(t1.Comp_Cost as decimal(18,2)) as comp_cost2, \'0\' as Cost_Increase, t1.contract_number, cast(t1.Comp_Cost as decimal(18,2)) as comp_cost3,  t1.\"TARGET_MARGIN(%)\", t1.\"CURRENT_MARGIN(%)\", cast(t1.SELL_PRICE as decimal(18,2)) as SELL_PRICE2, cast(t1.TARGET_PRICE as decimal(18,2)) as TARGET_PRICE2, t1.\"4% PRICE_MARGIN(%)\", cast(t1.\"4% PRICE\" as decimal(18,2)) as Price_4percent, cast(t1.\"ANN_4% PRICE_OPP\" as decimal(18,2)) as annual_opp_4percent, t1.weighted_cost FROM KGEK_MASTER_TABLE_V3 t1 inner join KGEK_NEG_MARGIN_BID_ITEM n1 on t1.prcsrc = n1.prcsrc and t1.RLITEM = n1.RLITEM where t1.bid_owner in(\'{owner_name1}\') and n1.neg_margin_bid_item_flag = 0 order by prcsrc desc, RLITEM desc, ACCT desc"
    tab3_query = f"select Group_no, Group_name, ACCT, account_name, count(*) as tot_cnt FROM KGEK_MASTER_TABLE_V3 where bid_owner in(\'{owner_name1}\') and sell_price >= weighted_cost and sell_price>=COMP_COST group by Group_no, Group_name, ACCT, account_name order by ACCT desc"
    # create tab1 input data list of lists
    negative_item_query = f"select t1.SVP_name, t1.Rep_mgr_name, t1.bid_owner, t1.prcsrc, t1.ACCT, t1.Account_name, t1.ITEM_DESCRIPTION, t1.MFG_CODE, t1.prod_cat_desc, t1.RLITEM, t1.QTY_6MO, cast(t1.SALES_6MO as decimal(18,2)) as sales_6mo6, t1.BHTYPE, cast(t1.Comp_Cost as decimal(18,2)) as comp_cost2, \'0\' as Cost_Increase, t1.contract_number, cast(t1.Comp_Cost as decimal(18,2)) as comp_cost3,  t1.\"TARGET_MARGIN(%)\", t1.\"CURRENT_MARGIN(%)\", cast(t1.SELL_PRICE as decimal(18,2)) as SELL_PRICE2, cast(t1.TARGET_PRICE as decimal(18,2)) as TARGET_PRICE2, t1.\"4% PRICE_MARGIN(%)\", cast(t1.\"4% PRICE\" as decimal(18,2)) as Price_4percent, cast(t1.\"ANN_4% PRICE_OPP\" as decimal(18,2)) as annual_opp_4percent, t1.weighted_cost FROM KGEK_MASTER_TABLE_V3 t1 inner join KGEK_NEG_MARGIN_BID_ITEM n1 on t1.prcsrc = n1.prcsrc and t1.RLITEM = n1.RLITEM where t1.bid_owner in(\'{owner_name1}\') and n1.neg_margin_bid_item_flag > 0 order by prcsrc desc, RLITEM desc, ACCT desc"
    cur.execute(tab1_query) # run the query
    tab1_all_rows = cur.fetchall() # get all rows for the query
    tab1_count_rows = len(tab1_all_rows)
    tab1_list_start = 0
    tab1_row_cell = 6
    # create tab2 input data list of lists
    cur.execute(tab2_query) # run the query
    tab2_all_rows = cur.fetchall() # get all rows for the query
    tab2_count_rows = len(tab2_all_rows)
    tab2_list_start = 0
    tab2_row_cell = 5
    # create tab3 input data list of lists
    cur.execute(tab3_query) # run the query
    tab3_all_rows = cur.fetchall() # get all rows for the query
    tab3_count_rows = len(tab3_all_rows)
    tab3_list_start = 0
    tab3_row_cell = 2

    cur.execute(negative_item_query) # run the query
    tabneg_all_rows = cur.fetchall() # get all rows for the query
    tabneg_count_rows = len(tabneg_all_rows)
    tabneg_list_start = 0
    tabneg_row_cell = 5

    # Insert tab1 data into tab1, one row at a time for the current bid owner
    while tab1_list_start < tab1_count_rows:
        current_row = tab1_all_rows[tab1_list_start]
        prcsrc, bhdesc, price_increase, bid_owner, ignore_tot_cnt = current_row
        GP_goal_dollars = "=SUMIF(\'Step2-Positive Margin Items\'!Price_Source,\'Step1-CustomerAction\'!A:A,Goal)+SUMIF(\'Step3-Negative Margin Items\'!Price_Source2,\'Step1-CustomerAction\'!A:A,Goal2)"
        opportunity_captured_dollars = "=SUMIF(\'Step2-Positive Margin Items\'!Price_Source,\'Step1-CustomerAction\'!A:A,Attainment)+SUMIF(\'Step3-Negative Margin Items\'!Price_Source2,\'Step1-CustomerAction\'!A:A,Attainment2)"
        #GP_goal_dollars = "=SUMIF(\'Step2-Positive Margin Items\'!Price_Source,\'Step1-CustomerAction\'!A:A,Goal)"
        #opportunity_captured_dollars = "=SUMIF(\'Step2-Positive Margin Items\'!Price_Source,\'Step1-CustomerAction\'!A:A,Attainment)"
        percent_attained = "=IF(ISERR(D:D/C:C) = FALSE, D:D/C:C, \"\")"
        Bid_Owner = "Yes"
        Ordering_System = "No"
        PSA_Account_Status = "No"
        Delay_Price_change = "No"
        tab1.cell(row = tab1_row_cell, column = 1, value = prcsrc)
        tab1.cell(row = tab1_row_cell, column = 2, value = bhdesc)
        tab1.cell(row = tab1_row_cell, column = 3, value = GP_goal_dollars)
        tab1.cell(row = tab1_row_cell, column = 4, value = opportunity_captured_dollars)
        tab1.cell(row = tab1_row_cell, column = 5, value = percent_attained)
        tab1.cell(row = tab1_row_cell, column = 6, value = price_increase)
        tab1.cell(row = tab1_row_cell, column = 7, value = Bid_Owner)
        tab1.cell(row = tab1_row_cell, column = 8, value = Ordering_System)
        tab1.cell(row = tab1_row_cell, column = 9, value = PSA_Account_Status)
        tab1.cell(row = tab1_row_cell, column = 10, value = Delay_Price_change)
        tab1_row_cell = tab1_row_cell + 1
        tab1_list_start = tab1_list_start + 1
    # Insert tab2 data into tab2, one row at a time for the current bid owner
    while tab2_list_start < tab2_count_rows:
        current_row = tab2_all_rows[tab2_list_start]
        leader_name, ASM, rep_name, prcsrc, customer_number, customer_name, item_description, MFG, product_desc, item_number, mth6_qty, mth6_sales, BHTYPE, current_cost, cost_increase, contract_item, new_cost, target_margin, current_margin, current_price, target_price, new_margin, new_price, gp_goal, weighted_cost  = current_row
        Add_Margin = "=If(Current< 0, \"Negative Margin\", Vlookup(Price_Source,\'Step1-CustomerAction\'!A:F,6, 0))"
        GP_Dollars_Captured = "=ROUND((Price-Current_Price)*Quantity*2,2)"
        Bid_Owner_2 = "= Vlookup(Price_Source,\'Step1-CustomerAction\'!A:G,7,FALSE)"
        PSA_2 = "=VLOOKUP(Price_source,'Step1-CustomerAction'!A:I,9,FALSE)"
        #GP_Dollars_Captured = "=(ROUND(New_Cost/(1-Total_Margin), 2)-ROUND(New_Cost/(1-Current),2)) *K:K *2"  # change formula
        tab2.cell(row = tab2_row_cell, column = 1, value = leader_name)
        tab2.cell(row = tab2_row_cell, column = 2, value = ASM)
        tab2.cell(row = tab2_row_cell, column = 3, value = rep_name)
        tab2.cell(row = tab2_row_cell, column = 4, value = prcsrc)
        tab2.cell(row = tab2_row_cell, column = 5, value = customer_number)
        tab2.cell(row = tab2_row_cell, column = 6, value = customer_name)
        tab2.cell(row = tab2_row_cell, column = 7, value = item_description)
        tab2.cell(row = tab2_row_cell, column = 8, value = MFG)
        tab2.cell(row = tab2_row_cell, column = 9, value = product_desc)
        tab2.cell(row = tab2_row_cell, column = 10, value = item_number)
        tab2.cell(row = tab2_row_cell, column = 11, value = mth6_qty)
        tab2.cell(row = tab2_row_cell, column = 12, value = mth6_sales)
        tab2.cell(row = tab2_row_cell, column = 13, value = BHTYPE)
        tab2.cell(row = tab2_row_cell, column = 14, value = current_cost)
        tab2.cell(row = tab2_row_cell, column = 15, value = weighted_cost)
        tab2.cell(row = tab2_row_cell, column = 16, value = contract_item)
        tab2.cell(row = tab2_row_cell, column = 17, value = new_cost)
        tab2.cell(row = tab2_row_cell, column = 18, value = target_margin)
        tab2.cell(row = tab2_row_cell, column = 19, value = current_margin)
        tab2.cell(row = tab2_row_cell, column = 20, value = current_price)
        tab2.cell(row = tab2_row_cell, column = 21, value = target_price)
        tab2.cell(row = tab2_row_cell, column = 22, value = Add_Margin)
        tab2.cell(row = tab2_row_cell, column = 23, value = new_margin)
        tab2.cell(row = tab2_row_cell, column = 24, value = new_price)
        tab2.cell(row = tab2_row_cell, column = 25, value = gp_goal)
        tab2.cell(row = tab2_row_cell, column = 26, value = GP_Dollars_Captured)
        tab2.cell(row = tab2_row_cell, column = 27, value = Bid_Owner_2)    # edited from 28 to 27 on 20171229
        tab2.cell(row = tab2_row_cell, column = 28, value = PSA_2)    # edited from 28 to 27 on 20171229
        tab2_row_cell = tab2_row_cell + 1
        tab2_list_start = tab2_list_start + 1


    # Insert tab2 data into tab2, one row at a time for the current bid owner
    while tabneg_list_start < tabneg_count_rows:
        current_row = tabneg_all_rows[tabneg_list_start]
        leader_name, ASM, rep_name, prcsrc, customer_number, customer_name, item_description, MFG, product_desc, item_number, mth6_qty, mth6_sales, BHTYPE, current_cost, cost_increase, contract_item, new_cost, target_margin, current_margin, current_price, target_price, new_margin, new_price, gp_goal, weighted_cost  = current_row
        Add_Margin = "=If(Current2< 0, \"Negative Margin\", Vlookup(Price_Source2,\'Step1-CustomerAction\'!A:F,6, 0))"
        GP_Dollars_Captured = "=ROUND((Price2-Current_Price2)*Quantity2*2,2)"
        Bid_Owner_2 = "= Vlookup(Price_Source2,\'Step1-CustomerAction\'!A:G,7,FALSE)"
        PSA_2 = "=VLOOKUP(Price_source2,'Step1-CustomerAction'!A:I,9,FALSE)"
        #GP_Dollars_Captured = "=(ROUND(New_Cost/(1-Total_Margin), 2)-ROUND(New_Cost/(1-Current),2)) *K:K *2"  # change formula
        tab_negative_margin_items.cell(row = tabneg_row_cell, column = 1, value = leader_name)
        tab_negative_margin_items.cell(row = tabneg_row_cell, column = 2, value = ASM)
        tab_negative_margin_items.cell(row = tabneg_row_cell, column = 3, value = rep_name)
        tab_negative_margin_items.cell(row = tabneg_row_cell, column = 4, value = prcsrc)
        tab_negative_margin_items.cell(row = tabneg_row_cell, column = 5, value = customer_number)
        tab_negative_margin_items.cell(row = tabneg_row_cell, column = 6, value = customer_name)
        tab_negative_margin_items.cell(row = tabneg_row_cell, column = 7, value = item_description)
        tab_negative_margin_items.cell(row = tabneg_row_cell, column = 8, value = MFG)
        tab_negative_margin_items.cell(row = tabneg_row_cell, column = 9, value = product_desc)
        tab_negative_margin_items.cell(row = tabneg_row_cell, column = 10, value = item_number)
        tab_negative_margin_items.cell(row = tabneg_row_cell, column = 11, value = mth6_qty)
        tab_negative_margin_items.cell(row = tabneg_row_cell, column = 12, value = mth6_sales)
        tab_negative_margin_items.cell(row = tabneg_row_cell, column = 13, value = BHTYPE)
        tab_negative_margin_items.cell(row = tabneg_row_cell, column = 14, value = current_cost)
        tab_negative_margin_items.cell(row = tabneg_row_cell, column = 15, value = weighted_cost)
        tab_negative_margin_items.cell(row = tabneg_row_cell, column = 16, value = contract_item)
        tab_negative_margin_items.cell(row = tabneg_row_cell, column = 17, value = new_cost)
        tab_negative_margin_items.cell(row = tabneg_row_cell, column = 18, value = target_margin)
        tab_negative_margin_items.cell(row = tabneg_row_cell, column = 19, value = current_margin)
        tab_negative_margin_items.cell(row = tabneg_row_cell, column = 20, value = current_price)
        tab_negative_margin_items.cell(row = tabneg_row_cell, column = 21, value = target_price)
        tab_negative_margin_items.cell(row = tabneg_row_cell, column = 22, value = Add_Margin)
        tab_negative_margin_items.cell(row = tabneg_row_cell, column = 23, value = new_margin)
        tab_negative_margin_items.cell(row = tabneg_row_cell, column = 24, value = new_price)
        tab_negative_margin_items.cell(row = tabneg_row_cell, column = 25, value = gp_goal)
        tab_negative_margin_items.cell(row = tabneg_row_cell, column = 26, value = GP_Dollars_Captured)
        tab_negative_margin_items.cell(row = tabneg_row_cell, column = 27, value = Bid_Owner_2)    # edited from 28 to 27 on 20171229
        tab_negative_margin_items.cell(row = tabneg_row_cell, column = 28, value = PSA_2)    # edited from 28 to 27 on 20171229
        tabneg_row_cell = tabneg_row_cell + 1
        tabneg_list_start = tabneg_list_start + 1


    # Insert tab3 data into tab3, one row at a time for the current bid owner
    while tab3_list_start < tab3_count_rows:
        current_row = tab3_all_rows[tab3_list_start]
        group_number, group_name, customer_number, customer_name, ignore_tot_cnt = current_row
        tab3.cell(row = tab3_row_cell, column = 1, value = group_number)
        tab3.cell(row = tab3_row_cell, column = 2, value = group_name)
        tab3.cell(row = tab3_row_cell, column = 3, value = customer_number)
        tab3.cell(row = tab3_row_cell, column = 4, value = customer_name)
        tab3_row_cell = tab3_row_cell + 1
        tab3_list_start = tab3_list_start + 1
    time = str(datetime.now())
    tab5.cell(row = 1, column = 4, value = time)
    tab5.cell(row = 1, column = 5, value = owner_name1)
    tab5.cell(row = 1, column = 6, value = bid_owner_email)



    #HF_rep_cnt = f"select count(*) from PRICING.EHVXZ0J_HF_FINAL_DRIVER_011218 where account_rep in(\'{current_bid_owner}\')"
    HF_rep_level_query = f"select * from PRICING.EHVXZ0J_HF_FINAL_DRIVER_011218 where account_rep in(\'{owner_name1}\')" # use bid_owner since Workbook is at rep-level
    cur.execute(HF_rep_level_query) # run the query
    HF_rep_in_check = cur.fetchall() # get all rows for the query
    tab4_count_rows = len(HF_rep_in_check)
    tab4_list_start = 0
    tab4_row_cell = 5

    if tab4_count_rows > 0:
        list_of_bid_owners_with_HF_tab.append(owner_name1)
        while tab4_list_start < tab4_count_rows:

            current_row = HF_rep_in_check[tab4_list_start]
            PSA_Account_Status_tab4 = "No"
            Account_Owner_Status_tab4 = "Yes"
            HF_Opt_Out_Status_tab4 = "No"

            ACCOUNT_REP, ACCOUNT_NUMBER, ACCOUNT_NAME, FY2017_SALES, GROSS_MARGIN, DELIVERED_GP, FY2017_ACTUAL_FREIGHT, FY2017_ACTUAL_HF_CHARGES, FY2017_HF_CHARGES, DEFAULT_PRICE_INCREASE, FREIGHT_CODE, HANDLING_FEE_CODE = current_row
            tab_HF.cell(row = tab4_row_cell, column = 1, value = ACCOUNT_NUMBER)
            tab_HF.cell(row = tab4_row_cell, column = 2, value = ACCOUNT_NAME)
            tab_HF.cell(row = tab4_row_cell, column = 3, value = HANDLING_FEE_CODE)
            tab_HF.cell(row = tab4_row_cell, column = 4, value = FY2017_SALES)
            tab_HF.cell(row = tab4_row_cell, column = 5, value = DELIVERED_GP)
            tab_HF.cell(row = tab4_row_cell, column = 6, value = GROSS_MARGIN)
            #tab_HF.cell(row = tab4_row_cell, column = 7, value = Customer_Charge_Freight)  # currently not in Craig's table
            tab_HF.cell(row = tab4_row_cell, column = 8, value = FY2017_ACTUAL_FREIGHT)
            tab_HF.cell(row = tab4_row_cell, column = 9, value = FY2017_HF_CHARGES)
            tab_HF.cell(row = tab4_row_cell, column = 10, value = FY2017_ACTUAL_HF_CHARGES)
            tab_HF.cell(row = tab4_row_cell, column = 11, value = DEFAULT_PRICE_INCREASE)
            tab_HF.cell(row = tab4_row_cell, column = 12, value = "=HF_Revenue * (1+HF_Rate_Increase)")

            tab_HF.cell(row = tab4_row_cell, column = 14, value = PSA_Account_Status_tab4)
            tab_HF.cell(row = tab4_row_cell, column = 15, value = Account_Owner_Status_tab4)
            tab_HF.cell(row = tab4_row_cell, column = 16, value = HF_Opt_Out_Status_tab4)
            tab4_row_cell += 1
            tab4_list_start += 1
    else:
        list_of_bid_owners_without_HF_tab.append(owner_name1)
        tab_HF.cell(row = tab4_row_cell, column = 2, value = "You have no accounts that are under review for a handling fee increase. Please move on to the next section of the Workbook")

    filename1 = f"C:\\Users\\ee1nj1k\\Documents\\Margin_Day\\201802\\WIP_Report\\Actual_Reports\\Margin_Day_Report_for_{owner_name1}_{current_date}.xlsm"
    filename_for_Excel = f"C:\\\\Users\\\\ee1nj1k\\\\Documents\\\\Margin_Day\\\\201802\\\\WIP_Report\\\\Actual_Reports\\\\Margin_Day_Report_for_{owner_name1}_{current_date}.xlsm"
    wb.save(filename = filename1)

    # run macro to change formatting
    excel=win32com.client.Dispatch("Excel.Application")
    wb2 = excel.Workbooks.Open(Filename = filename1)
    excel.Application.Run("StupidExportFixinator")
    excel.Visible = True
    wb2.Close(True) # True = saves the file when closing it
    excel.Quit()

    current_file_and_email_list = []
    current_file_and_email_list.append(filename_for_Excel)
    current_file_and_email_list.append(bid_owner_email)
    current_file_and_email_list.append(owner_name1)

    list_of_file_locations.append(current_file_and_email_list)
    # save emails with attachment as draft (number of attachments per minute ~4 to 5)
    print(owner_name1)
    current_bid_owner += 1

cur.close()
con.close()

Excel_File_Locations = Workbook()
ws1 = Excel_File_Locations.active

current_row = 0
number_of_files = len(list_of_file_locations)
row_cell = 2
ws1.cell(row = 1, column = 1, value = "File_Location")
ws1.cell(row = 1, column = 2, value = "Email_Address")
ws1.cell(row = 1, column = 3, value = "Bid_Owner")

while current_row < number_of_files:
    file_location1, bid_owner_email1, owner_name11 = list_of_file_locations[current_row]
    ws1.cell(row = row_cell, column = 1, value = file_location1)
    ws1.cell(row = row_cell, column = 2, value = bid_owner_email1)
    ws1.cell(row = row_cell, column = 3, value = owner_name11)
    row_cell = row_cell + 1
    current_row = current_row + 1

Excel_File_Locations.save("C:\\Users\\ee1nj1k\\Documents\\Margin_Day\\201802\\WIP_Report\\Actual_Reports\\List_of_File_Locations_and_Emails_20171229.xlsx" )


print("complete")

print(list_of_bid_owners_with_HF_tab)
print("without")
print(list_of_bid_owners_without_HF_tab)

#print(list_of_file_locations)

exit()


    #excel.Workbooks.Open(Filename = filename1, ReadOnly=1)
    #Define_Module = "StupidExportFixinator"
    #Define_Module = f"Template_20170523_{owner_name1}.xlsm!module1.StupidExportFixinator"
    #excel.Application.Run("StupidExportFixinator")
    #excel.Application.Save() # if you want to save then uncomment this line and change delete the ", ReadOnly=1" part from the open function.
    #excel.Application.Quit() # Comment this out if your excel script closes
    #wb2.Application.Save() # if you want to save then uncomment this line and change delete the ", ReadOnly=1" part from the open function.
    #wb2.Application.Quit() # Comment this out if your excel script closes
    #wb2.Application.save()
